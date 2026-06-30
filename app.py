

from flask import Flask, jsonify, request, render_template, send_file
from datetime import datetime, timedelta
from pathlib import Path
import re, os, threading, uuid, ssl, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False

# ── Paths ────────────────────────────────────────────────────────────────────
BASE       = Path(__file__).parent
EXCEL_PATH = BASE / "data" / "library_records.xlsx"
EXCEL_PATH.parent.mkdir(exist_ok=True)

# ── Email — read from env vars, fall back to defaults you can edit ────────────
SMTP_HOST     = "smtp.gmail.com"
SMTP_PORT     = 587
SMTP_USER     = os.environ.get("SMTP_USER", "Your-Mail@gmail.com")      # ← edit
SMTP_PASS     = os.environ.get("SMTP_PASS", "xxxx xxxx xxxx xxxx")    # ← edit
LIBRARY_NAME  = "Citadel"
LIBRARY_ADDR  = "The Citadel, Honeywine River District, Oldtown, The Reach, Westeros"
LIBRARIAN_TEL = "Send Raven to Tower 4"
LIBRARY_EMAIL = "hello@citadel.in"

# ── Books ────────────────────────────────────────────────────────────────────
from data.books import BOOKS

# ── In-memory stores ─────────────────────────────────────────────────────────
_rentals         = []        # rental records
_card_members    = {}        # card_no → member dict
_record_counter  = [0]
_card_counter    = [1000]
_excel_lock      = threading.Lock()


def _next_rid():
    _record_counter[0] += 1
    return f"REC{_record_counter[0]:05d}"

def _next_card():
    _card_counter[0] += 1
    return f"CLB-{datetime.now().year}-{_card_counter[0]}"


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL  (real SMTP — no silent swallowing)
# ═══════════════════════════════════════════════════════════════════════════════
def _send(to_email: str, subject: str, html: str) -> tuple[bool, str]:
    """Send one HTML email. Returns (success, error_message)."""
    try:
        msg               = MIMEMultipart("alternative")
        msg["Subject"]    = subject
        msg["From"]       = f"{LIBRARY_NAME} <{SMTP_USER}>"
        msg["To"]         = to_email
        msg["Reply-To"]   = LIBRARY_EMAIL
        msg.attach(MIMEText(html, "html", "utf-8"))

        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, to_email, msg.as_string())
        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, "SMTP authentication failed. Check SMTP_USER and SMTP_PASS."
    except smtplib.SMTPException as e:
        return False, f"SMTP error: {e}"
    except Exception as e:
        return False, f"Email error: {e}"


def _card_email_html(name: str, card_no: str) -> str:
    return f"""
<html><body style="font-family:Arial,sans-serif;background:#F5F0E8;margin:0;padding:0">
<div style="max-width:600px;margin:30px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)">
  <div style="background:#1A1209;padding:32px 36px">
    <h1 style="font-family:Georgia,serif;color:#C9A84C;margin:0;font-size:26px">Citadel</h1>
    <p style="color:rgba(245,240,232,.5);font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:4px 0 0">Library Card Issued</p>
  </div>
  <div style="padding:32px 36px">
    <p style="color:#3A2E1E;font-size:16px">Dear <strong>{name}</strong>,</p>
    <p style="color:#5A4A3A;line-height:1.7">Welcome to Citadel! Your library card has been registered. Here are your details:</p>
    <div style="background:#F5F0E8;border-left:4px solid #C9A84C;border-radius:4px;padding:24px 28px;margin:24px 0;text-align:center">
      <p style="margin:0 0 6px;color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:2px;text-transform:uppercase">Your Library Card Number</p>
      <p style="margin:0;font-family:monospace;font-size:28px;font-weight:bold;color:#1A1209;letter-spacing:4px">{card_no}</p>
      <p style="margin:10px 0 0;font-size:12px;color:#9A8A7A">Keep this number safe — you'll need it to rent books</p>
    </div>
    <div style="background:#FFF8E7;border:1px solid #C9A84C;border-radius:4px;padding:16px 20px;margin:20px 0">
      <p style="margin:0;color:#7B4F2E;font-weight:bold;font-size:14px">📍 Collect your physical card from:</p>
      <p style="margin:6px 0 0;color:#5A4A3A">{LIBRARY_ADDR}</p>
      <p style="margin:4px 0 0;color:#5A4A3A">📞 {LIBRARIAN_TEL}</p>
      <p style="margin:8px 0 0;color:#7B4F2E;font-size:13px"><strong>Hours:</strong> Mon–Fri 9 AM–8 PM &nbsp;|&nbsp; Sat–Sun 10 AM–6 PM</p>
    </div>
    <p style="color:#5A4A3A;line-height:1.7;font-size:14px">With your card you can borrow any available book for <strong>30 days</strong>, completely free. Happy reading!</p>
  </div>
  <div style="background:#EDE5D4;padding:16px 36px;text-align:center">
    <p style="margin:0;color:#7B4F2E;font-size:12px">&copy; 2026 {LIBRARY_NAME} · {LIBRARY_ADDR}</p>
  </div>
</div></body></html>"""


def _rental_email_html(name, book, card_no, rent_date, due_date) -> str:
    return f"""
<html><body style="font-family:Arial,sans-serif;background:#F5F0E8;margin:0;padding:0">
<div style="max-width:600px;margin:30px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)">
  <div style="background:#1A1209;padding:32px 36px">
    <h1 style="font-family:Georgia,serif;color:#C9A84C;margin:0;font-size:26px">Citadel</h1>
    <p style="color:rgba(245,240,232,.5);font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:4px 0 0">Rental Confirmation</p>
  </div>
  <div style="padding:32px 36px">
    <p style="color:#3A2E1E;font-size:16px">Dear <strong>{name}</strong>,</p>
    <p style="color:#5A4A3A;line-height:1.7">Your book rental has been confirmed. Please collect your book from the library counter and present your Library Card.</p>
    <div style="background:#F5F0E8;border-left:4px solid #C9A84C;border-radius:4px;padding:20px 24px;margin:24px 0">
      <table style="width:100%;border-collapse:collapse">
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0;width:140px">Book</td><td style="color:#1A1209;font-weight:bold;font-size:15px">{book['title']}</td></tr>
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0">Author</td><td style="color:#3A2E1E">{book['author']}</td></tr>
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0">Genre</td><td style="color:#3A2E1E">{book['genre']}</td></tr>
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0">Card No.</td><td style="color:#3A2E1E;font-family:monospace">{card_no}</td></tr>
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0">Rented On</td><td style="color:#3A2E1E">{rent_date}</td></tr>
        <tr><td style="color:#7B4F2E;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;padding:5px 0">Due Date</td><td style="color:#C0392B;font-weight:bold;font-size:15px">{due_date}</td></tr>
      </table>
    </div>
    <div style="background:#FFF8E7;border:1px solid #C9A84C;border-radius:4px;padding:16px 20px;margin:20px 0">
      <p style="margin:0;color:#7B4F2E;font-weight:bold">📍 Collect your book from:</p>
      <p style="margin:6px 0 0;color:#5A4A3A">{LIBRARY_ADDR}</p>
      <p style="margin:4px 0 0;color:#5A4A3A">📞 {LIBRARIAN_TEL}</p>
      <p style="margin:8px 0 0;color:#7B4F2E;font-size:13px"><strong>Hours:</strong> Mon–Fri 9 AM–8 PM &nbsp;|&nbsp; Sat–Sun 10 AM–6 PM</p>
    </div>
    <p style="color:#5A4A3A;line-height:1.7;font-size:14px">Please return by <strong style="color:#C0392B">{due_date}</strong> to avoid late fees. Happy reading!</p>
  </div>
  <div style="background:#EDE5D4;padding:16px 36px;text-align:center">
    <p style="margin:0;color:#7B4F2E;font-size:12px">&copy; 2026 {LIBRARY_NAME}</p>
  </div>
</div></body></html>"""


def _return_email_html(name, title, return_date) -> str:
    return f"""
<html><body style="font-family:Arial,sans-serif;background:#F5F0E8;margin:0;padding:0">
<div style="max-width:600px;margin:30px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)">
  <div style="background:#1A1209;padding:32px 36px">
    <h1 style="font-family:Georgia,serif;color:#C9A84C;margin:0;font-size:26px">Citadel</h1>
    <p style="color:rgba(245,240,232,.5);font-size:12px;letter-spacing:2px;text-transform:uppercase;margin:4px 0 0">Return Confirmation</p>
  </div>
  <div style="padding:32px 36px">
    <p style="color:#3A2E1E;font-size:16px">Dear <strong>{name}</strong>,</p>
    <p style="color:#5A4A3A;line-height:1.7">Thank you for returning <strong>"{title}"</strong> on {return_date}. We hope you enjoyed reading it!</p>
    <p style="color:#5A4A3A;line-height:1.7;margin-top:1rem">Visit us any time to borrow your next great read. Our full catalogue is available at the library and online.</p>
  </div>
  <div style="background:#EDE5D4;padding:16px 36px;text-align:center">
    <p style="margin:0;color:#7B4F2E;font-size:12px">&copy; 2026 {LIBRARY_NAME}</p>
  </div>
</div></body></html>"""


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
HEADERS = ["Record ID","Member Name","Email","Library Card No.",
           "Book ID","Book Title","Author","Genre",
           "Rent Date","Due Date","Return Date","Status","Notes"]

def _thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def init_excel():
    if EXCEL_PATH.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental Records"
    hf = PatternFill("solid", fgColor="1A1209")
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(1, i, h)
        c.font      = Font(name="Arial", bold=True, color="C9A84C", size=11)
        c.fill      = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin()
    for i, w in enumerate([12,22,30,18,8,35,25,18,14,14,14,12,30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Citadel – Rental Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A1209")
    for r, lbl, formula in [
        (3,"Total Records","=COUNTA('Rental Records'!A2:A50000)"),
        (4,"Currently Rented","=COUNTIF('Rental Records'!L2:L50000,\"Rented\")"),
        (5,"Returned","=COUNTIF('Rental Records'!L2:L50000,\"Returned\")"),
    ]:
        ws2.cell(r,1,lbl).font = Font(name="Arial", bold=True)
        ws2.cell(r,2,formula).font = Font(name="Arial", color="7B4F2E")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    wb.save(EXCEL_PATH)

def append_rental_excel(rec):
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Rental Records"]
        row = ws.max_row + 1
        vals = [rec["record_id"],rec["member_name"],rec["email"],rec["card_no"],
                rec["book_id"],rec["book_title"],rec["author"],rec["genre"],
                rec["rent_date"],rec["due_date"],rec.get("return_date",""),
                rec["status"],rec.get("notes","")]
        af = PatternFill("solid", fgColor="FDF8EE") if row % 2 == 0 else None
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            c.font      = Font(name="Arial", size=10)
            c.border    = _thin()
            c.alignment = Alignment(vertical="center", wrap_text=(i==13))
            if af: c.fill = af
            if i == 12:
                col = {"Rented":"27AE60","Returned":"7B4F2E","Overdue":"C0392B"}.get(v,"000000")
                c.font = Font(name="Arial", size=10, bold=True, color=col)
        wb.save(EXCEL_PATH)

def mark_returned_excel(record_id, return_date):
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Rental Records"]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(record_id):
                row[10].value = return_date
                row[11].value = "Returned"
                row[11].font  = Font(name="Arial", size=10, bold=True, color="7B4F2E")
                break
        wb.save(EXCEL_PATH)

# Also track card registrations in a second sheet
def append_card_excel(card_no, name, email, date):
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Card Members" not in wb.sheetnames:
            ws2 = wb.create_sheet("Card Members")
            hf  = PatternFill("solid", fgColor="1A1209")
            for i, h in enumerate(["Card No.","Name","Email","Issued Date"], 1):
                c = ws2.cell(1, i, h)
                c.font      = Font(name="Arial", bold=True, color="C9A84C", size=11)
                c.fill      = hf
                c.alignment = Alignment(horizontal="center")
                c.border    = _thin()
            ws2.column_dimensions["A"].width = 18
            ws2.column_dimensions["B"].width = 24
            ws2.column_dimensions["C"].width = 32
            ws2.column_dimensions["D"].width = 16
        else:
            ws2 = wb["Card Members"]
        r = ws2.max_row + 1
        for i, v in enumerate([card_no, name, email, date], 1):
            ws2.cell(r, i, v).font = Font(name="Arial", size=10)
        wb.save(EXCEL_PATH)

init_excel()

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Pages
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/book/<int:book_id>")
def book_detail(book_id):
    book = next((b for b in BOOKS if b["id"] == book_id), None)
    if not book:
        return "Book not found", 404
    return render_template("book.html", book=book)

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Books API
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/books")
def api_books():
    genre  = request.args.get("genre","")
    search = request.args.get("search","").lower()
    sort   = request.args.get("sort","")
    result = list(BOOKS)
    if genre and genre != "All":
        result = [b for b in result if b["genre"] == genre]
    if search:
        result = [b for b in result if search in b["title"].lower() or search in b["author"].lower()]
    if sort == "title":        result.sort(key=lambda b: b["title"])
    elif sort == "author":     result.sort(key=lambda b: b["author"])
    elif sort == "newest":     result.sort(key=lambda b: b.get("year",0), reverse=True)
    elif sort == "price_asc":  result.sort(key=lambda b: b["price"])
    elif sort == "price_desc": result.sort(key=lambda b: b["price"], reverse=True)
    return jsonify(result)

@app.route("/api/books/new")
def api_new_books():
    nb = [b for b in BOOKS if b.get("new")]
    nb.sort(key=lambda b: b.get("year",0), reverse=True)
    return jsonify(nb[:10])

@app.route("/api/books/<int:book_id>")
def api_book(book_id):
    book = next((b for b in BOOKS if b["id"] == book_id), None)
    return jsonify(book) if book else (jsonify({"error":"Not found"}), 404)

@app.route("/api/genres")
def api_genres():
    return jsonify(["All"] + sorted(set(b["genre"] for b in BOOKS)))

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Library Card Registration
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/get-card", methods=["POST"])
def api_get_card():
    data  = request.get_json()
    name  = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    phone = (data.get("phone") or "").strip()

    if not name or not email:
        return jsonify({"error": "Name and email are required."}), 400
    if not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email):
        return jsonify({"error": "Please enter a valid email address."}), 400
    # check duplicate
    if email in [m["email"] for m in _card_members.values()]:
        existing = next(m for m in _card_members.values() if m["email"] == email)
        return jsonify({"error": f'A card for this email already exists: {existing["card_no"]}'}), 409

    card_no   = _next_card()
    issued    = datetime.now().strftime("%d %b %Y")
    member    = {"card_no": card_no, "name": name, "email": email,
                 "phone": phone, "issued": issued}
    _card_members[card_no] = member

    # Excel (background)
    threading.Thread(target=append_card_excel,
                     args=(card_no, name, email, issued), daemon=True).start()

    # Email — send synchronously so we can report real errors to client
    ok, err = _send(
        email,
        f"🎉 Your Citadel Card — {card_no}",
        _card_email_html(name, card_no)
    )
    email_status = "sent" if ok else f"failed: {err}"

    return jsonify({
        "card_no":      card_no,
        "name":         name,
        "email":        email,
        "issued":       issued,
        "email_status": email_status,
        "message":      f"Your library card {card_no} has been issued!"
    }), 201

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Rent
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/rent", methods=["POST"])
def api_rent():
    data    = request.get_json()
    name    = (data.get("name") or "").strip()
    email   = (data.get("email") or "").strip().lower()
    card_no = (data.get("card_no") or "").strip()
    book_id = int(data.get("book_id", 0))

    if not name or not email or not card_no:
        return jsonify({"error": "Name, email and library card number are required."}), 400
    if not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email):
        return jsonify({"error": "Please enter a valid email address."}), 400
    if len(card_no) < 3:
        return jsonify({"error": "Please enter a valid library card number."}), 400

    book = next((b for b in BOOKS if b["id"] == book_id), None)
    if not book:
        return jsonify({"error": "Book not found."}), 404
    if not book["available"]:
        return jsonify({"error": f'"{book["title"]}" is currently checked out.'}), 409
    if next((r for r in _rentals if r["card_no"]==card_no and r["book_id"]==book_id and r["status"]=="Rented"), None):
        return jsonify({"error": "This card already has this book rented."}), 409

    rent_date = datetime.now().strftime("%d %b %Y")
    due_date  = (datetime.now() + timedelta(days=30)).strftime("%d %b %Y")
    record_id = _next_rid()
    rec = {
        "record_id": record_id, "member_name": name, "email": email,
        "card_no": card_no, "book_id": book_id, "book_title": book["title"],
        "author": book["author"], "genre": book["genre"],
        "rent_date": rent_date, "due_date": due_date,
        "return_date": "", "status": "Rented",
        "notes": f"Collect from library counter. Card: {card_no}",
    }
    _rentals.append(rec)
    book["available"] = False

    threading.Thread(target=append_rental_excel, args=(rec,), daemon=True).start()

    # Send email — synchronous so client gets real status
    ok, err = _send(
        email,
        f"📚 Book Rented: {book['title']} – {LIBRARY_NAME}",
        _rental_email_html(name, book, card_no, rent_date, due_date)
    )
    return jsonify({
        "message":      f'"{book["title"]}" rented successfully!',
        "record_id":    record_id,
        "due_date":     due_date,
        "rent_date":    rent_date,
        "email_status": "sent" if ok else f"queued (email error: {err})",
    }), 201

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Return
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/return", methods=["POST"])
def api_return():
    data    = request.get_json()
    card_no = (data.get("card_no") or "").strip()
    book_id = int(data.get("book_id", 0))

    rental = next((r for r in _rentals if r["card_no"]==card_no and r["book_id"]==book_id and r["status"]=="Rented"), None)
    if not rental:
        return jsonify({"error": "No active rental found for this card and book."}), 404

    return_date           = datetime.now().strftime("%d %b %Y")
    rental["status"]      = "Returned"
    rental["return_date"] = return_date

    book = next((b for b in BOOKS if b["id"] == book_id), None)
    if book: book["available"] = True

    threading.Thread(target=mark_returned_excel, args=(rental["record_id"], return_date), daemon=True).start()

    ok, _ = _send(
        rental["email"],
        f"✅ Book Returned: {rental['book_title']} – {LIBRARY_NAME}",
        _return_email_html(rental["member_name"], rental["book_title"], return_date)
    )
    return jsonify({"message": f'"{rental["book_title"]}" returned. Thank you!'})

@app.route("/api/my-rentals")
def api_my_rentals():
    card_no = request.args.get("card_no","").strip()
    if not card_no: return jsonify([])
    return jsonify([r for r in _rentals if r["card_no"]==card_no and r["status"]=="Rented"])

if __name__ == "__main__":
    app.run(debug=True, port=5000)
